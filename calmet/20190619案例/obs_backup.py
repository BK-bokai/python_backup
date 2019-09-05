import datetime
import time
import numpy as np
import pandas as pd
import csv
from openpyxl import load_workbook
from openpyxl import Workbook


start='2019-06-18-00:00'
end=  '2019-06-21-00:00'
time_key={}
title=['時間','WS','WD','ICELL','ICC','AT','RH','P','天氣型態']
time_key['titie']=','.join(title)
datestart=datetime.datetime.strptime(start,'%Y-%m-%d-%H:%M')
dateend=datetime.datetime.strptime(end,'%Y-%m-%d-%H:%M')
yk = pd.read_excel("善化20190618-20.xlsx")


# print(len(yk['時間']))

for i in range(0,len(yk['時間'])):
        print (yk['時間'][i])


while(datestart < dateend):
        time_key[datestart]=(','.join([str("9999"),str("9999"),str("9999"),str("9999"),str("9999"),str("9999"),str("9999"),str("9999")]))
        for i in range(0,len(yk['時間'])):
                if (datestart==datetime.datetime.strptime(yk['時間'][i],'%Y-%m-%d-%H')):
                        data=[str(yk['WS'][i]),str(yk['WD'][i]),str(yk['ICELL'][i]),str(yk['ICC'][i]),str(yk['AT'][i]),str(yk['RH'][i]),str(yk['P'][i]),str(yk['天氣型態'][i])]
                        data= ','.join(data)
                        time_key[datestart]=data

                
                        

        datestart+=datetime.timedelta(minutes=10)
        
# for key, value in time_key.items() :
#         print (key)
#         print (value)



wb = Workbook()
ws = wb.active
# ws2 = wb.create_sheet('永康')
ws.title = "善化"
# ws.append([time_key['title']])
for key, value in time_key.items() :
        # writer.writerow([key,str(value)])
        ws.append([key,str(value)])
# 儲存成 create_sample.xlsx 檔案
wb.save('善化.xlsx')


# ele_array = np.zeros( (len(hrlen),17))